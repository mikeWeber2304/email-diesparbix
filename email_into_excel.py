import imaplib
import email
import re
import datetime
import openpyxl
import html2text
from email.mime.multipart import MIMEMultipart
import smtplib
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from openpyxl.styles import colors
from openpyxl.styles import Font
from statistics import mode
from progress.bar import FillingSquaresBar
from progress.spinner import PieSpinner
from termcolor import colored
import time
import sys


# to login
def update_time():
    time_rn = str(datetime.datetime.now().strftime("%d.%m.%y %H:%M"))
    return time_rn

current_time = update_time()
postbox_date = input('Bitte geben Sie das zu bearbeitende Postfach an: ')
print('\nIhre Eingabe lautete: ' + '\"' + postbox_date + '\"')

if re.match('^[0-9]{2}_[0-9]{2}$', postbox_date):
    print('Bitte warten.')
#elif re.match('^$', postbox_date):
    #print('Da Sie keine Eingabe gemacht haben, wird der normale Posteingang ausgewählt.')
    #postbox_date = 'INBOX'
else:
    print(
        'Bitte machen Sie eine korrekte Eingabe nachdem Sie das Programm erneut ausgeführt haben. Das Programm wird jetzt beendet.')
    sys.exit()

hostname = 'imap.gmail.com'
sendserver = smtplib.SMTP('smtp.gmail.com', 587)
username = 'test3528040@gmail.com'
password = 'LabraLenny98'

#logging in
con = imaplib.IMAP4_SSL(hostname)
con.login(username, password)
con.select(postbox_date)

#global variables
BigBuyList = []
AmazonList = []
allMails = []
all_lists_AZ = []
all_lists_BB = []
matching_lists = []
OtherList = []
all_sorted_lists = []
specialCases = []
filepath_base = "/home/mike/Schreibtisch/E-Mail-Verwaltung-DieSparbix/" + current_time + ".xlsx"
special_cases_sorted = []
no_match = []



def get_emails(result_bytes):
    msgs = []
    for num in result_bytes[0].split():
        typ, data = con.fetch(num, '(RFC822)')
        msgs.append(data)
    return msgs


def search(key, value, con):
    result, data = con.search(None, key, '"{}"'.format(value))
    return data


def get_body(mail):
    if mail.is_multipart():
        return get_body(mail.get_payload(0))
    else:
        return mail.get_payload(None, True)

try:
    result, data = con.search(None, "ALL")
    inbox_item_list = data[0].split()
    print("Postfach valide. Bitte warten")
except:
    print('Das von Ihnen angegebene Postfach existiert nicht. Bitte geben Sie ein valides Postfach an, das Programm wird jetzt beendet.')
    sys.exit()


def delete_all_mails():
    print('\n')
    time.sleep(0.1)
    bar = FillingSquaresBar('Mails werden entfernt', max=len(inbox_item_list))
    bar.start()
    for num in inbox_item_list:
        con.store(num, '+FLAGS', '\\Deleted')
        bar.next()
    con.expunge()
    bar.finish()
    con.close()
    con.logout()

def get_all_mails():
    bar = FillingSquaresBar('Mails werden verarbeitet ', max=len(inbox_item_list))
    bar.start()
    for item in inbox_item_list:
        result2, email_data = con.fetch(item, '(RFC822)')
        raw = email_data[0][1]
        email_message = email.message_from_bytes(raw)
        allMails.append(email_message)  # allMails is now a list packed up with all mails
        bar.next()
    bar.finish()

def get_msg_date():
    all_times = []
    mail_time = ''
    for mail in allMails:
        all_times.append(mail['Date'][5:16])
    t1 = mode(all_times)
    t2 = t1.replace(' ', '_')
    return t2

def sort_all_mails():
    bar = FillingSquaresBar('Mails werden sortiert', max=allMails.__len__())
    bar.start()
    for item in allMails:
        if "BigBuy" in item['Subject']:
            BigBuyList.append(item)
        if "Artikel" in item['Subject']:
            AmazonList.append(item)
        else:
            OtherList.append(item)
        bar.next()
    bar.finish()

def get_info_on_order_AZ(a):  # a is a mail parsed to a string by getBody, spits out the important information the mail in a list
    infolistAZ = []
    a1 = a.split('\\')
    a2 = ''.join(a1)
    a3 = a2.split('rn')
    a4 = ''.join(a3)
    a5 = a4.split(',')
    a6 = ''.join(a5)
    infolistAZ.append(re.search('Bestellnummer: (.+?)Bitte', a6).group(1))
    infolistAZ.append(re.search('SKU: (.+?)Menge:', a6).group(1))
    infolistAZ.append(re.search('Artikel: (.+?)Zustand:', a6).group(1))
    a7 = re.search('Preis: (.+?)rn', a2).group(1)
    a8 = a7.strip('EUR ')
    a9 = ''.join(a8)
    if '.' in a9:
        a10 = a9.replace('.', '')
    else:
        a10 = a9
    a11 = ''.join(a10)
    a12 = a11.replace(',', '.')
    infolistAZ.append(float(a12))
    a13 = re.search('Versandkosten: (.+?)rn', a2).group(1)
    a14 = a13.strip('\'EUR ')
    a15 = ''.join(a14)
    if '.' in a9:
        a16 = a15.replace('.', '')
    else:
        a16 = a15
    a17 = ''.join(a16)
    a18 = a17.replace(',', '.')
    infolistAZ.append(float(a18))
    preisGesamt = float(a12) + float(a18)
    infolistAZ.append(preisGesamt)
    a19 = re.search('Amazon-Gebxc3xbchren: \((.+?)\)', a2).group(1)
    a20 = a19.strip('\'EUR ')
    a21 = ''.join(a20)
    if '.' in a9:
        a22 = a21.strip('.')
    else:
        a22 = a21
    a23 = ''.join(a22)
    a24 = a23.replace(',', '.')
    infolistAZ.append(float(a24))

    return infolistAZ


def get_info_on_order_BB(a):

    infolistBB = []
    atest = html2text.html2text(a)
    a2 = ''.join(atest)
    a3 = a2.split('\\')
    a4 = ''.join(a3)
    a5 = a4.split('|')
    a6 = ''.join(a5)
    a7 = a6.split('rn')
    a8 = ''.join(a7)
    a9 = a4.splitlines()
    a10 = ''.join(a9)
    infolistBB.append(re.search('## SHIPPING INFORMATIONrn  ---  (.+?)rn', a10).group(1))
    infolistBB.append(re.search('Order no\.\*\*     (.+?) ', a8).group(1))
    a11 = re.search('Total sum\*\*\|  (.+?)xc', a4).group(1)
    if ',' in a11:
        a12 = a11.replace(',', '')
    else:
        a12 = a11
    infolistBB.append(float(a12))
    if re.search('Inteal ref\.:\*\*     (.+?) ', a8) is None:
        infolistBB.append('Bestellung manuell getaetigt.')
    else:
        infolistBB.append(re.search('Inteal ref\.:\*\*     (.+?) ', a8).group(1))


    return infolistBB


def get_all_info_BB():
    bar = FillingSquaresBar('Infos von BigBuy werden verarbeitet', max=BigBuyList.__len__())
    bar.start()
    for item in BigBuyList:
        all_lists_BB.append(get_info_on_order_BB(str(get_body(item))))
        bar.next()
    bar.finish()

def get_all_info_AZ():
    bar = FillingSquaresBar('Infos von Amazon werden verarbeitet', max=AmazonList.__len__())
    bar.start()
    for item in AmazonList:
        all_lists_AZ.append(get_info_on_order_AZ(str(get_body(item))))
        bar.next()
    bar.finish()


def check_BB_and_AZ_orders():
    bar = FillingSquaresBar('Bestellnummern werden abgeglichen', max=all_lists_BB.__len__())
    bar.start()
    for item in all_lists_BB:
        for item2 in all_lists_AZ:
            if item[3] == 'Bestellung manuell getaetigt.':  # this is done 26 times, why?
                specialCases.append(item)
            if item[3] == item2[0]:
                item2.extend(item)
                matching_lists.append(item2)
        bar.next()
    bar.finish()
    return matching_lists


def remove_duplicates(list):
    list_duplicates_removed = []
    for num in list:
        if num not in list_duplicates_removed:
            list_duplicates_removed.append(num)
    for item in list_duplicates_removed:
        del item[2]
    return list_duplicates_removed

def write_in_excel():
    spinner = PieSpinner('Excel-Dokument wird erstellt ')
    spinner.start()
    ft_amazon = Font(color=colors.BLUE)
    ft_negative = Font(color=colors.RED)
    ft_positive = Font(color=colors.GREEN)
    ft_magenta = Font(color=colors.DARKYELLOW)

    headlines = ['ID Amazon', 'ID BigBuy', 'Kunde', 'VK Netto', 'Versand Netto', 'Gesamt', 'Amazon', 'BigBuy', 'Ertrag',
                 'SKU', 'Produkt']
    wb = openpyxl.Workbook()
    ws = wb.active
    spinner.next()
    ws.cell(row=30, column=4, value="=SUM(D2:D29)")
    ws.cell(row=30, column=5, value="=SUM(E2:E29)")
    ws.cell(row=30, column=6, value="=SUM(F2:F29)")
    ws.cell(row=30, column=7, value="=SUM(G2:G29)")
    ws.cell(row=30, column=8, value="=SUM(H2:H29)")
    ws.cell(row=30, column=9, value="=SUM(I2:I29)")
    spinner.next()
    for zeile in range(1, 12):
        ws.cell(row=1, column=zeile, value=headlines[zeile - 1])
        spinner.next()

    for row in range(2, all_sorted_lists.__len__() + 2):
        spinner.next()
        for column in range(1, 12):
            ws.cell(row=row, column=column, value=all_sorted_lists[row - 2][column - 1])

    fancy_list = remove_duplicates(no_match)

    ws.cell(row=32, column=1, value='KEINE ZUORDNUNG GEFUNDEN').font = ft_negative

    for row in range(fancy_list.__len__()):
        spinner.next()
        for column in range(1, 4):
            ws.cell(row=33+row, column=column, value=fancy_list[row-1][column-1])

    for i in range(2,31):
        ws.cell(row=i, column=8).font = ft_magenta

    for i in range(2,31):
        ws.cell(row=i, column=7).font = ft_amazon

    for i in range(2, 32):
        spinner.next()
        currentCell = ws.cell(row=i, column=9)
        if '-' in str(currentCell.value):
            currentCell.font = ft_negative
        else:
            currentCell.font = ft_positive

    def get_font_of_cell(row, column):
        cell = ws.cell(row=row, column=column)
        if '-' in str(cell.value):
            cell.font = ft_negative
        else:
            cell.font = ft_positive

    get_font_of_cell(30, 9)

    wb.save(filepath_base)
    spinner.finish()

def pop_all_matching_lists():
    for item in matching_lists:
        item.pop()

def make_the_lists_alrighty():
    for item in matching_lists:
            earnings = item[5] - float(item[6]) - float(item[9])
            sortedlist = [item[0], item[8], item[7], item[3], item[4], item[5], item[6], item[9], earnings, item[1], item[2]]
            all_sorted_lists.append(sortedlist)
    return all_sorted_lists


def backcheck_matches():
    valid_match = []
    for item in all_lists_BB:
        for item2 in all_sorted_lists:
            if item[3] == item2[0]:
                item.append('NOICE')
    for item in all_lists_BB:
        if 'NOICE' in item:
            valid_match.append(item)
        else:
            no_match.append(item)
    return no_match

def send_mail_back():

    spinner = PieSpinner('\nMail wird versendet ')
    spinner.start()

    msg = MIMEMultipart()
    body = "Das von Ihnen angeforderte Excel-Dokument wurde erstellt, Sie finden es in den Anhaengen." \
           "\nErstellzeitpunkt: " + current_time + \
           "\nTagesumsatz vom: " + get_msg_date() + \
           "\nAnzahl verarbeiteter Mails: " +str(len(inbox_item_list)) + \
           "\nAnzahl verarbeiteter Mails von Amazon: " +str(AmazonList.__len__()) + \
           "\nAnzahl verarbeiteter Mails von BigBuy: " +str(BigBuyList.__len__()) + \
           "\nAnzahl nicht zugeordneter Bestellungen: " +str(no_match.__len__()) + \
           "\ndavon manuell getätigt: " +str(remove_duplicates(specialCases).__len__())

    msg['From'] = username
    msg['To'] = "info@diesparbix.com"
    msg['Subject'] = "Excel-Dokument erstellt."

    spinner.next()

    msg.attach(MIMEText(body, 'plain'))
    time.sleep(5)
    anhang = open(filepath_base, 'rb')

    spinner.next()

    part = MIMEBase('application', 'octet-stream')
    part.set_payload(anhang.read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', "attachement, filename = " + get_msg_date() + '.xlsx')

    spinner.next()

    msg.attach(part)
    text = msg.as_string()

    spinner.next()

    server = smtplib.SMTP('smtp.gmail.com', 587)
    server.connect("smtp.gmail.com", 587)
    server.ehlo()
    server.starttls()
    server.login(username, password)

    spinner.next()

    server.sendmail(msg['From'], msg['To'], text)
    server.quit()

    spinner.finish()


def main():
    try:
        get_all_mails()
    except ZeroDivisionError:
        print('Der gewählte Posteingang ist leer, bitte stellen Sie sicher, dass der von Ihnen gewählte Posteingang nicht leer ist. Das Programm wird jetzt beendet.')
        sys.exit()
    sort_all_mails()
    get_all_info_AZ()
    get_all_info_BB()
    check_BB_and_AZ_orders()
    pop_all_matching_lists()
    make_the_lists_alrighty()
    backcheck_matches()
    write_in_excel()
    send_mail_back()
    if postbox_date == 'INBOX':
        delete_all_mails()
    print(colored('\nProgramm beendet', 'green'))

main()
